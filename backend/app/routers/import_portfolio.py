"""
Portfolio import endpoint.

Accepts xlsx files (B3 CEI export) or raw text and uses Claude Haiku
to extract structured portfolio positions.
"""

import io
import json
import logging
from typing import Optional

import anthropic
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from pydantic import BaseModel

from app.config import settings
from app.dependencies_mobile import verify_mobile_api_key

logger = logging.getLogger(__name__)

router = APIRouter(
    prefix="/api/mobile/import",
    tags=["import"],
    dependencies=[Depends(verify_mobile_api_key)],
)

SYSTEM_PROMPT = """You are a financial data extraction assistant. Extract all investment positions from the provided portfolio data.

Return a JSON object with a "positions" array where each item has these exact fields:
- "ticker": string — the trading code (e.g. "ITUB3", "BTLG11", "AAPL"). For B3 assets, use the Codigo de Negociacao. For CDBs/renda fixa, use a descriptive code like "CDB-C6-2028" with issuer and maturity year.
- "display_name": string — full company/fund name
- "quantity": number — number of shares/units/cotas
- "current_price": number — latest price per unit
- "asset_class": string — one of: "acoesBR", "fiis", "usStocks", "reits", "crypto", "rendaFixa". Classify BDRs as "usStocks". Classify FIIs (Fundo de Investimento Imobiliario) as "fiis". Classify CDBs/LCIs/LCAs/Tesouro as "rendaFixa".
- "total_value": number — total position value

Rules:
- Skip empty rows, totals, headers, and summary lines
- For quantity, use the "Quantidade Disponivel" or "Quantidade" field
- For price, use "Preco de Fechamento" or "Preço Atualizado" or similar
- Clean ticker names: remove trailing spaces, use just the trading code"""

POSITIONS_SCHEMA = {
    "type": "object",
    "properties": {
        "positions": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "ticker": {"type": "string"},
                    "display_name": {"type": "string"},
                    "quantity": {"type": "number"},
                    "current_price": {"type": "number"},
                    "asset_class": {
                        "type": "string",
                        "enum": [
                            "acoesBR",
                            "fiis",
                            "usStocks",
                            "reits",
                            "crypto",
                            "rendaFixa",
                        ],
                    },
                    "total_value": {"type": "number"},
                },
                "required": ["ticker"],
                "additionalProperties": False,
            },
        }
    },
    "required": ["positions"],
    "additionalProperties": False,
}


class ParsedPosition(BaseModel):
    ticker: str
    display_name: str = ""
    quantity: float = 0
    current_price: float = 0
    asset_class: str = "acoesBR"
    total_value: float = 0


class ImportResponse(BaseModel):
    positions: list[ParsedPosition]
    raw_text_preview: str = ""


def _extract_text_from_xlsx(file_bytes: bytes) -> str:
    """Extract text content from xlsx file for LLM processing."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl not installed on server. Install with: pip install openpyxl",
        )

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    lines = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"\n=== {sheet_name} ===")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):  # skip fully empty rows
                lines.append("\t".join(cells))

    wb.close()
    return "\n".join(lines)


def _extract_text_from_csv(file_bytes: bytes) -> str:
    """Extract text from CSV file."""
    for encoding in ("utf-8", "latin-1", "cp1252"):
        try:
            return file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="replace")


async def _call_claude(text: str) -> list[dict]:
    """Send portfolio text to Claude Haiku and parse the structured response."""
    if not settings.anthropic_api_key:
        raise HTTPException(status_code=500, detail="Anthropic API key not configured")

    client = anthropic.AsyncAnthropic(api_key=settings.anthropic_api_key)

    try:
        response = await client.messages.create(
            model="claude-haiku-4-5",
            max_tokens=16000,
            system=SYSTEM_PROMPT,
            messages=[
                {
                    "role": "user",
                    "content": f"Extract positions from this portfolio data:\n\n{text}",
                }
            ],
            output_config={
                "format": {
                    "type": "json_schema",
                    "schema": POSITIONS_SCHEMA,
                }
            },
        )
    except anthropic.APIStatusError as e:
        logger.error("Claude API error: %s %s", e.status_code, str(e)[:500])
        raise HTTPException(status_code=502, detail="Claude API error")
    except anthropic.APIConnectionError:
        logger.error("Claude API connection error")
        raise HTTPException(status_code=502, detail="Could not reach Claude API")

    text_block = next(
        (b.text for b in response.content if b.type == "text"), None
    )
    if not text_block:
        logger.error("No text block in Claude response: %s", response.content)
        raise HTTPException(status_code=502, detail="Unexpected Claude response")

    try:
        data = json.loads(text_block)
        return data.get("positions", [])
    except json.JSONDecodeError:
        logger.error("Failed to parse Claude JSON: %s", text_block[:500])
        raise HTTPException(status_code=502, detail="Failed to parse AI response as JSON")


@router.post("/parse", response_model=ImportResponse)
async def parse_portfolio(
    file: Optional[UploadFile] = File(None),
    text: Optional[str] = Form(None),
):
    """
    Parse a portfolio export file (xlsx/csv) or raw text using Claude AI.
    Returns structured positions ready for import.
    """
    if file is not None:
        file_bytes = await file.read()
        if len(file_bytes) > 5 * 1024 * 1024:  # 5 MB limit
            raise HTTPException(status_code=413, detail="File too large. Maximum size is 5 MB.")
        filename = (file.filename or "").lower()

        if filename.endswith(".xlsx"):
            extracted = _extract_text_from_xlsx(file_bytes)
        elif filename.endswith(".csv") or filename.endswith(".txt"):
            extracted = _extract_text_from_csv(file_bytes)
        else:
            raise HTTPException(status_code=400, detail="Unsupported file type. Use .xlsx, .csv, or .txt")
    elif text:
        extracted = text
    else:
        raise HTTPException(status_code=400, detail="Provide either a file or text")

    # Truncate if too long (context limit)
    if len(extracted) > 50_000:
        extracted = extracted[:50_000]

    positions_raw = await _call_claude(extracted)

    positions = []
    for p in positions_raw:
        try:
            positions.append(
                ParsedPosition(
                    ticker=p.get("ticker", "").strip(),
                    display_name=p.get("display_name", "").strip(),
                    quantity=float(p.get("quantity", 0)),
                    current_price=float(p.get("current_price", 0)),
                    asset_class=p.get("asset_class", "acoesBR"),
                    total_value=float(p.get("total_value", 0)),
                )
            )
        except (ValueError, TypeError):
            continue  # skip malformed entries

    preview = extracted[:500] + ("..." if len(extracted) > 500 else "")

    return ImportResponse(positions=positions, raw_text_preview=preview)
